#!/usr/bin/env python3
"""
Working Excel Test Plan Generator

Creates clean Excel file that opens without corruption.
All worksheets fully populated with real test data.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from pathlib import Path


def create_working_excel():
    """Create working Excel file."""
    
    print("🔧 Creating Working Excel Test Plan")
    print("=" * 50)
    
    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Styles
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    critical_fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
    critical_font = Font(color="FFFFFF", bold=True)
    
    # 1. Setup Instructions
    print("   1. Setup Instructions")
    sheet1 = wb.create_sheet("Setup Instructions")
    create_setup_data(sheet1, header_fill, header_font)
    
    # 2. Migration Tests (CRITICAL)
    print("   2. Migration Tests")
    sheet2 = wb.create_sheet("Migration Tests") 
    create_migration_data(sheet2, header_fill, header_font, critical_fill, critical_font)
    
    # 3. API Authentication
    print("   3. API Authentication")
    sheet3 = wb.create_sheet("API Authentication")
    create_auth_data(sheet3, header_fill, header_font)
    
    # 4. API Teams
    print("   4. API Teams")
    sheet4 = wb.create_sheet("API Teams")
    create_teams_data(sheet4, header_fill, header_font)
    
    # 5. API Servers
    print("   5. API Servers") 
    sheet5 = wb.create_sheet("API Servers")
    create_servers_data(sheet5, header_fill, header_font)
    
    # 6. Admin UI
    print("   6. Admin UI")
    sheet6 = wb.create_sheet("Admin UI")
    create_ui_data(sheet6, header_fill, header_font)
    
    # 7. Database Tests
    print("   7. Database Tests")
    sheet7 = wb.create_sheet("Database Tests")
    create_db_data(sheet7, header_fill, header_font)
    
    # 8. Security Tests
    print("   8. Security Tests")
    sheet8 = wb.create_sheet("Security Tests")
    create_security_data(sheet8, header_fill, header_font, critical_fill, critical_font)
    
    # Save file properly
    filepath = Path("test-plan.xlsx")
    print("\\n💾 Saving file...")
    
    wb.save(filepath)
    print("✅ File saved")
    
    # CRITICAL: Close workbook properly
    wb.close()
    print("✅ File closed")
    
    # Verify
    print("\\n🔍 Verifying...")
    try:
        test_wb = openpyxl.load_workbook(filepath)
        print(f"✅ Opens successfully: {len(test_wb.worksheets)} worksheets")
        
        # Check key worksheets
        for sheet in test_wb.worksheets:
            test_count = max(0, sheet.max_row - 1)
            print(f"   📄 {sheet.title}: {test_count} tests")
        
        test_wb.close()
        print("✅ Test file closed")
        
        print("\\n🎊 SUCCESS! Working Excel file created!")
        return True
        
    except Exception as e:
        print(f"❌ Verification failed: {e}")
        return False


def create_setup_data(sheet, header_fill, header_font):
    """Create setup instructions data."""
    
    headers = ["Step", "Action", "Command", "Expected", "Status", "Notes"]
    for i, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=i, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    data = [
        ["1", "Check Prerequisites", "python3 --version && git --version", "Python 3.11+ and Git installed", "☐", "Must have both"],
        ["2", "Clone Repository", "git clone <repository-url>", "Repository downloaded", "☐", "Get URL from admin"],
        ["3", "Enter Directory", "cd mcp-context-forge", "Directory changed", "☐", ""],
        ["4", "Copy Environment", "cp .env.example .env", "Environment file created", "☐", ""],
        ["5", "Edit Configuration", "vi .env", "Configuration edited", "☐", "Set admin email/password"],
        ["6", "Install Dependencies", "make install-dev", "Dependencies installed", "☐", "May take 5-10 minutes"],
        ["7", "Run Migration", "python3 -m mcpgateway.bootstrap_db", "Migration completed", "☐", "CRITICAL STEP"],
        ["8", "Verify Migration", "python3 scripts/verify_multitenancy_0_7_0_migration.py", "All checks pass", "☐", "Must pass"],
        ["9", "Start Gateway", "make dev", "Server running on port 4444", "☐", "Keep terminal open"],
        ["10", "Test Health", "curl http://localhost:4444/health", '{"status":"ok"}', "☐", "Basic connectivity"],
        ["11", "Access Admin UI", "Open http://localhost:4444/admin", "Login page loads", "☐", ""],
        ["12", "Test Login", "Login with admin credentials", "Dashboard appears", "☐", "Main validation"]
    ]
    
    for i, row in enumerate(data, 2):
        for j, value in enumerate(row, 1):
            sheet.cell(row=i, column=j, value=value)
    
    auto_size_columns(sheet)


def create_migration_data(sheet, header_fill, header_font, critical_fill, critical_font):
    """Create migration test data."""
    
    headers = ["Test ID", "Priority", "Component", "Description", "Steps", "Expected", "Actual", "Status", "Tester", "Comments", "SQLite", "PostgreSQL"]
    
    for i, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=i, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    data = [
        ["MIG-001", "CRITICAL", "Admin User", "Platform admin created", "Check admin user exists in database", "Admin user found with is_admin=true", "", "☐", "", "", "✓", "✓"],
        ["MIG-002", "CRITICAL", "Personal Team", "Admin personal team exists", "Run verification script", "Personal team found", "", "☐", "", "", "✓", "✓"],
        ["MIG-003", "CRITICAL", "Server Visibility", "OLD SERVERS VISIBLE - MAIN TEST", "Open admin UI, navigate to Virtual Servers", "ALL servers visible including old ones", "", "☐", "", "MAIN MIGRATION TEST", "✓", "✓"],
        ["MIG-004", "CRITICAL", "Resource Teams", "Resources assigned to teams", "Check team assignments in UI and DB", "All resources have team_id populated", "", "☐", "", "", "✓", "✓"],
        ["MIG-005", "CRITICAL", "Email Auth", "Email authentication works", "Test login with email/password", "Email login successful", "", "☐", "", "", "✓", "✓"],
        ["MIG-006", "HIGH", "Basic Auth", "Basic auth compatibility", "Test basic authentication", "Basic auth still works", "", "☐", "", "", "✓", "✓"],
        ["MIG-007", "HIGH", "API Functionality", "APIs respond correctly", "Test core API endpoints", "All APIs return expected responses", "", "☐", "", "", "✓", "✓"],
        ["MIG-008", "MEDIUM", "Team Membership", "Admin team ownership", "Check admin is team owner", "Admin listed as owner of personal team", "", "☐", "", "", "✓", "✓"]
    ]
    
    for i, row in enumerate(data, 2):
        for j, value in enumerate(row, 1):
            cell = sheet.cell(row=i, column=j, value=value)
            if j == 2 and value == "CRITICAL":  # Priority column
                cell.fill = critical_fill
                cell.font = critical_font
    
    auto_size_columns(sheet)


def create_auth_data(sheet, header_fill, header_font):
    """Create authentication API data."""
    
    headers = ["Test ID", "Endpoint", "Method", "Description", "cURL Command", "Expected Status", "Expected Response", "Actual Status", "Actual Response", "Status", "Tester", "Comments"]
    
    for i, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=i, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    data = [
        ["AUTH-001", "/auth/register", "POST", "User registration", 'curl -X POST http://localhost:4444/auth/register -d \'{"email":"test@example.com","password":"Test123"}\'', "201", "User created", "", "", "☐", "", ""],
        ["AUTH-002", "/auth/login", "POST", "Email login", 'curl -X POST http://localhost:4444/auth/login -d \'{"email":"admin@example.com","password":"changeme"}\'', "200", "JWT token", "", "", "☐", "", ""],
        ["AUTH-003", "/auth/logout", "POST", "User logout", 'curl -X POST http://localhost:4444/auth/logout -H "Authorization: Bearer <TOKEN>"', "200", "Logout success", "", "", "☐", "", ""],
        ["AUTH-004", "/auth/refresh", "POST", "Token refresh", 'curl -X POST http://localhost:4444/auth/refresh -H "Authorization: Bearer <TOKEN>"', "200", "New token", "", "", "☐", "", ""],
        ["AUTH-005", "/auth/profile", "GET", "User profile", 'curl http://localhost:4444/auth/profile -H "Authorization: Bearer <TOKEN>"', "200", "Profile data", "", "", "☐", "", ""],
        ["AUTH-006", "/auth/sso/github", "GET", "GitHub SSO", 'curl -I http://localhost:4444/auth/sso/github', "302", "GitHub redirect", "", "", "☐", "", ""],
        ["AUTH-007", "/auth/sso/google", "GET", "Google SSO", 'curl -I http://localhost:4444/auth/sso/google', "302", "Google redirect", "", "", "☐", "", ""],
        ["AUTH-008", "/auth/change-password", "POST", "Password change", 'curl -X POST http://localhost:4444/auth/change-password -H "Authorization: Bearer <TOKEN>" -d password_data', "200", "Password updated", "", "", "☐", "", ""]
    ]
    
    for i, row in enumerate(data, 2):
        for j, value in enumerate(row, 1):
            sheet.cell(row=i, column=j, value=value)
    
    auto_size_columns(sheet)


def create_teams_data(sheet, header_fill, header_font):
    """Create teams API data."""
    
    headers = ["Test ID", "Endpoint", "Method", "Description", "cURL Command", "Expected Status", "Expected Response", "Status", "Tester", "Comments"]
    
    for i, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=i, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    data = [
        ["TEAM-001", "/teams", "GET", "List teams", 'curl http://localhost:4444/teams -H "Authorization: Bearer <TOKEN>"', "200", "Team array", "☐", "", ""],
        ["TEAM-002", "/teams", "POST", "Create team", 'curl -X POST http://localhost:4444/teams -d team_data -H "Authorization: Bearer <TOKEN>"', "201", "Team created", "☐", "", ""],
        ["TEAM-003", "/teams/{id}", "GET", "Team details", 'curl http://localhost:4444/teams/{ID} -H "Authorization: Bearer <TOKEN>"', "200", "Team details", "☐", "", ""],
        ["TEAM-004", "/teams/{id}/members", "GET", "Team members", 'curl http://localhost:4444/teams/{ID}/members -H "Authorization: Bearer <TOKEN>"', "200", "Member list", "☐", "", ""],
        ["TEAM-005", "/teams/{id}/invitations", "POST", "Create invitation", 'curl -X POST http://localhost:4444/teams/{ID}/invitations -d invite_data', "201", "Invitation sent", "☐", "", ""]
    ]
    
    for i, row in enumerate(data, 2):
        for j, value in enumerate(row, 1):
            sheet.cell(row=i, column=j, value=value)
    
    auto_size_columns(sheet)


def create_servers_data(sheet, header_fill, header_font):
    """Create servers API data."""
    
    headers = ["Test ID", "Endpoint", "Method", "Description", "cURL Command", "Expected Status", "Status", "Tester"]
    
    for i, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=i, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    data = [
        ["SRV-001", "/servers", "GET", "List servers", 'curl http://localhost:4444/servers -H "Authorization: Bearer <TOKEN>"', "200", "☐", ""],
        ["SRV-002", "/servers", "POST", "Create server", 'curl -X POST http://localhost:4444/servers -d server_data', "201", "☐", ""],
        ["SRV-003", "/servers/{id}", "GET", "Server details", 'curl http://localhost:4444/servers/{ID} -H "Authorization: Bearer <TOKEN>"', "200", "☐", ""],
        ["SRV-004", "/servers/{id}/sse", "GET", "SSE connection", 'curl -N http://localhost:4444/servers/{ID}/sse -H "Authorization: Bearer <TOKEN>"', "200", "☐", ""],
        ["SRV-005", "/servers/{id}/tools", "GET", "Server tools", 'curl http://localhost:4444/servers/{ID}/tools -H "Authorization: Bearer <TOKEN>"', "200", "☐", ""]
    ]
    
    for i, row in enumerate(data, 2):
        for j, value in enumerate(row, 1):
            sheet.cell(row=i, column=j, value=value)
    
    auto_size_columns(sheet)


def create_ui_data(sheet, header_fill, header_font):
    """Create UI test data."""
    
    headers = ["Test ID", "Component", "Action", "Steps", "Expected", "Status", "Tester", "Browser", "Screenshot"]
    
    for i, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=i, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    data = [
        ["UI-001", "Login", "Test login", "Login with admin creds", "Dashboard loads", "☐", "", "Chrome", "Optional"],
        ["UI-002", "Dashboard", "View dashboard", "Check stats and navigation", "Dashboard functional", "☐", "", "Chrome", "Optional"],
        ["UI-003", "Servers", "View servers - MAIN TEST", "Navigate to Virtual Servers", "ALL SERVERS VISIBLE", "☐", "", "Chrome", "REQUIRED"],
        ["UI-004", "Teams", "Team management", "Navigate to Teams", "Teams functional", "☐", "", "Chrome", "Optional"],
        ["UI-005", "Tools", "Tool interface", "View and invoke tools", "Tools accessible", "☐", "", "Chrome", "Optional"],
        ["UI-006", "Export", "Config export", "Export configuration", "Export works", "☐", "", "Chrome", "Recommended"]
    ]
    
    for i, row in enumerate(data, 2):
        for j, value in enumerate(row, 1):
            sheet.cell(row=i, column=j, value=value)
    
    auto_size_columns(sheet)


def create_db_data(sheet, header_fill, header_font):
    """Create database test data."""
    
    headers = ["Test ID", "Database", "Feature", "Command", "Expected", "Status", "Performance", "Notes"]
    
    for i, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=i, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    data = [
        ["DB-001", "SQLite", "Migration", "python3 -m mcpgateway.bootstrap_db", "Success", "☐", "Fast", ""],
        ["DB-002", "SQLite", "Data Check", "sqlite3 mcp.db 'SELECT COUNT(*) FROM servers;'", "Count > 0", "☐", "Fast", ""],
        ["DB-003", "PostgreSQL", "Migration", "Set PG URL, run migration", "Success", "☐", "Fast", ""],
        ["DB-004", "PostgreSQL", "Advanced Types", "Test UUID, JSONB", "Advanced features work", "☐", "Excellent", ""],
        ["DB-005", "Both", "Performance", "Large dataset test", "Good performance", "☐", "Variable", ""]
    ]
    
    for i, row in enumerate(data, 2):
        for j, value in enumerate(row, 1):
            sheet.cell(row=i, column=j, value=value)
    
    auto_size_columns(sheet)


def create_security_data(sheet, header_fill, header_font, critical_fill, critical_font):
    """Create security test data."""
    
    headers = ["Test ID", "Attack Type", "Target", "Description", "Expected Defense", "Risk Level", "Status", "Tester", "Notes"]
    
    for i, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=i, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    data = [
        ["SEC-001", "SQL Injection", "API", "SQL injection attempt", "Input sanitized", "Critical", "☐", "", ""],
        ["SEC-002", "JWT Tampering", "Auth", "Token manipulation", "Token rejected", "Critical", "☐", "", ""],
        ["SEC-003", "Team Bypass", "Authorization", "Cross-team access", "Access denied", "Critical", "☐", "", ""],
        ["SEC-004", "XSS Attack", "UI", "Script injection", "Scripts escaped", "High", "☐", "", ""],
        ["SEC-005", "Brute Force", "Login", "Password attack", "Account locked", "Medium", "☐", "", ""]
    ]
    
    for i, row in enumerate(data, 2):
        for j, value in enumerate(row, 1):
            cell = sheet.cell(row=i, column=j, value=value)
            if j == 6 and value == "Critical":  # Risk Level
                cell.fill = critical_fill
                cell.font = critical_font
    
    auto_size_columns(sheet)


def auto_size_columns(sheet):
    """Auto-size columns."""
    
    for col in range(1, sheet.max_column + 1):
        max_length = 0
        for row in range(1, min(sheet.max_row + 1, 20)):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        
        width = min(max(max_length + 2, 10), 50)
        sheet.column_dimensions[get_column_letter(col)].width = width


if __name__ == "__main__":
    try:
        success = create_working_excel()
        if not success:
            sys.exit(1)
    except Exception as e:
        print(f"❌ Error: {e}")
        sys.exit(1)